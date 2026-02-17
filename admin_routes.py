from fastapi import (
    APIRouter,
    HTTPException,
    status,
    Depends,
    Query
)
from pydantic import BaseModel, EmailStr
from datetime import timedelta, datetime
from typing import Optional, List
import logging
import io

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from auth import (
    authenticate_admin,
    create_access_token,
    verify_token,
    ACCESS_TOKEN_EXPIRE_MINUTES
)

admin_router = APIRouter(prefix="/api/admin", tags=["Admin"])
logger = logging.getLogger(__name__)

# =====================================================
# AUTH MODELS
# =====================================================
class AdminLoginRequest(BaseModel):
    email: EmailStr
    password: str


class AdminLoginResponse(BaseModel):
    access_token: str
    token_type: str
    email: str
    expires_in: int


class StatusUpdate(BaseModel):
    status: str


# =====================================================
# LOGIN
# =====================================================
@admin_router.post("/login", response_model=AdminLoginResponse)
def admin_login(payload: AdminLoginRequest):
    """
    Admin login using email & password
    """
    try:
        user = authenticate_admin(payload.email, payload.password)

        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid email or password"
            )

        expires_delta = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)

        access_token = create_access_token(
            data={
                "sub": user["email"],
                "role": user.get("role", "admin")
            },
            expires_delta=expires_delta
        )

        return AdminLoginResponse(
            access_token=access_token,
            token_type="bearer",
            email=user["email"],
            expires_in=ACCESS_TOKEN_EXPIRE_MINUTES * 60
        )

    except HTTPException:
        raise
    except Exception:
        logger.exception("ADMIN LOGIN FAILED")
        raise HTTPException(
            status_code=500,
            detail="Internal Server Error"
        )


# =====================================================
# COVERAGE INQUIRIES (ADMIN)
# =====================================================
@admin_router.get("/coverage-inquiries")
async def admin_coverage_inquiries(
    current_user: dict = Depends(verify_token),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    limit: int = Query(100, le=500)
):
    from server import db

    query = {}

    if status and status != "all":
        query["status"] = status

    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"city": {"$regex": search, "$options": "i"}}
        ]

    inquiries = await db.coverage_inquiries.find(query).sort(
        "timestamp", -1
    ).limit(limit).to_list(limit)

    for i in inquiries:
        i["_id"] = str(i["_id"])

    return {"count": len(inquiries), "inquiries": inquiries}


# =====================================================
# PACKAGE INQUIRIES (ADMIN)
# =====================================================
@admin_router.get("/package-inquiries")
async def admin_package_inquiries(
    current_user: dict = Depends(verify_token),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    limit: int = Query(100, le=500)
):
    from server import db

    query = {}

    if status and status != "all":
        query["status"] = status

    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"inquiry_number": {"$regex": search, "$options": "i"}}
        ]

    inquiries = await db.package_inquiries.find(query).sort(
        "timestamp", -1
    ).limit(limit).to_list(limit)

    for i in inquiries:
        i["_id"] = str(i["_id"])

    return {"count": len(inquiries), "inquiries": inquiries}


# =====================================================
# UPDATE STATUS
# =====================================================
@admin_router.patch("/coverage-inquiry/{inquiry_id}/status")
async def update_coverage_status(
    inquiry_id: str,
    payload: StatusUpdate,
    current_user: dict = Depends(verify_token)
):
    from server import db

    result = await db.coverage_inquiries.update_one(
        {"id": inquiry_id},
        {"$set": {"status": payload.status}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inquiry not found")

    return {"message": "Status updated"}


@admin_router.patch("/package-inquiry/{inquiry_id}/status")
async def update_package_status(
    inquiry_id: str,
    payload: StatusUpdate,
    current_user: dict = Depends(verify_token)
):
    from server import db

    result = await db.package_inquiries.update_one(
        {"id": inquiry_id},
        {"$set": {"status": payload.status}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inquiry not found")

    return {"message": "Status updated"}


# =====================================================
# SETTINGS (EDITABLE)
# =====================================================
@admin_router.put("/settings/{setting_type}")
async def update_settings(
    setting_type: str,
    payload: dict,
    current_user: dict = Depends(verify_token)
):
    from server import db

    await db.site_settings.update_one(
        {"type": setting_type},
        {"$set": payload},
        upsert=True
    )

    return {"message": "Settings updated successfully"}


# =====================================================
# EXPORT EXCEL
# =====================================================
def create_excel(data: List[dict], headers: List[str]):
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in data:
        ws.append([row.get(h.lower(), "") for h in headers])

    return wb


@admin_router.get("/export/coverage")
async def export_coverage(current_user: dict = Depends(verify_token)):
    from server import db

    data = await db.coverage_inquiries.find().to_list(1000)
    wb = create_excel(data, ["ID", "Name", "Phone", "City", "Status", "Timestamp"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=coverage_{datetime.now():%Y%m%d}.xlsx"
        }
    )


@admin_router.get("/export/package")
async def export_package(current_user: dict = Depends(verify_token)):
    from server import db

    data = await db.package_inquiries.find().to_list(1000)
    wb = create_excel(data, ["Inquiry_number", "Name", "Email", "Phone", "Status", "Timestamp"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=package_{datetime.now():%Y%m%d}.xlsx"
        }
    )
